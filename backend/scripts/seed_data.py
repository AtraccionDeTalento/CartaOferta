"""
Seed data — importa datos reales de los Excel al sistema.
Ejecutar: python -m backend.scripts.seed_data
"""
import os
import sys
from datetime import datetime

# Setup path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from backend.database import engine, SessionLocal, init_db
from backend.models import (
    UnidadNegocio, Categoria, CategoriaTrabajador, TipoDocumento,
    Puesto, Sede, Correlativo, SolicitudIngreso, Movimiento,
)


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date'):
        return val
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _to_float(val):
    if val is None:
        return None
    val_str = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not val_str:
        return None
    try:
        return float(val_str)
    except ValueError:
        return None



def seed_catalogos(db):
    """Insertar catálogos base."""
    # Unidades de negocio
    for nombre in ["USIL", "CSIR", "IE", "EPG", "GEF"]:
        if not db.query(UnidadNegocio).filter(UnidadNegocio.nombre == nombre).first():
            db.add(UnidadNegocio(nombre=nombre))

    # Categorías
    for nombre in ["ACADEMICOS", "ADMINISTRATIVOS", "OTROS ACADEMICOS", "ADMINISTRATIVO"]:
        if not db.query(Categoria).filter(Categoria.nombre == nombre).first():
            db.add(Categoria(nombre=nombre))

    # Categorías trabajador
    for nombre in [
        "DIRECCION", "CONFIANZA - NO SUJETO A FISCALIZACION",
        "CONFIANZA - SUJETO A FISCALIZACION",
        "NINGUNA - NO SUJETO A FISCALIZACION",
        "NINGUNA - SUJETO A FISCALIZACION",
        "Sujeto a Fiscalización",
    ]:
        if not db.query(CategoriaTrabajador).filter(CategoriaTrabajador.nombre == nombre).first():
            db.add(CategoriaTrabajador(nombre=nombre))

    # Tipos de documento (movimientos)
    tipos = [
        ("Bono por funciones adicionales", None),
        ("Bono por excelencia", None),
        ("Bono por interinazgo", None),
        ("Convenio de Encargatura", None),
        ("Asignación de Movilidad", None),
        ("Convenio de Transporte", None),
        ("UPGRADE", None),
        ("Promoción", None),
        ("Ajuste salarial", None),
        ("Incremento", None),
        ("Incremento por retención", None),
        ("Carta de Movimiento", "Cambio de area"),
        ("Carta de Movimiento", "Cambio de jefe"),
        ("Carta de Movimiento", "Cambio de Sede"),
        ("Carta de Movimiento", "Cambio de Sucursal"),
        ("Carta de Movimiento", "Cambio de puesto"),
        ("Cese de convenio", "Encargatura"),
        ("Cese de convenio", "Bono transporte"),
        ("Cese de convenio", "Bono funciones adicionales"),
        ("Cese de convenio", "Bono Movilidad"),
        ("Cambio de Condición Laboral", None),
        ("Convenio de Cooperación Academica", None),
        ("Carta de Nombramiento", None),
        ("Bono por ingreso", None),
        ("Cambio de jefe", None),
        ("Cambio de puesto", None),
    ]
    for nombre, subtipo in tipos:
        exists = db.query(TipoDocumento).filter(
            TipoDocumento.nombre == nombre,
            TipoDocumento.subtipo == subtipo
        ).first()
        if not exists:
            db.add(TipoDocumento(nombre=nombre, subtipo=subtipo))

    # Sedes
    for nombre in ["UNIVERSIDAD", "CSIR", "ESCUELA DE POSTGRADO",
                    "Instituto de Emprendedores", "Toulón 348", "Lima Norte"]:
        if not db.query(Sede).filter(Sede.nombre == nombre).first():
            db.add(Sede(nombre=nombre))

    # Correlativos iniciales
    for tipo in ["solicitud_ingreso", "carta_oferta"]:
        if not db.query(Correlativo).filter(Correlativo.tipo == tipo).first():
            db.add(Correlativo(tipo=tipo, ultimo_valor=2409))

    db.commit()
    print("  Catalogos insertados OK")


def seed_from_excel(db):
    """Importar datos reales de los Excel."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl no disponible, saltando import Excel")
        return

    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    parent = os.path.dirname(base)

    # ── SOLICITUDES DE INGRESO ──
    excel1 = os.path.join(parent, "SOLICITUD DE CARTAS OFERTA_PROYECTO.xlsx")
    if os.path.exists(excel1):
        print(f"  Importando solicitudes desde: {excel1}")
        wb = openpyxl.load_workbook(excel1, data_only=True)
        ws = wb["Mar-A la fecha"]
        count = 0
        puestos_set = set()

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=29, values_only=False):
            vals = [cell.value for cell in row]
            nombre = vals[7]  # Col H = NOMBRES Y APELLIDOS
            if not nombre or str(nombre).strip() == "":
                continue

            nombre = str(nombre).strip()
            puesto_val = str(vals[9] or "").strip()  # Col J = PUESTO
            ceco_val = str(vals[10] or "").strip()   # Col K = CODIGO CR

            # Recopilar puestos únicos
            if puesto_val:
                puestos_set.add((puesto_val, ceco_val))

            s = SolicitudIngreso(
                correlativo=vals[0],  # Col A
                unidad=str(vals[1] or "").strip(),  # Col B
                fecha_solicitud=_to_date(vals[2]),  # Col C
                fecha_tentativa_ingreso=_to_date(vals[3]),  # Col D
                modalidad=str(vals[4] or "").strip(),  # Col E
                genero=str(vals[5] or "").strip(),  # Col F
                dni=str(vals[6] or "").strip(),  # Col G
                nombres_apellidos=nombre,
                puesto_nuevo=str(vals[8] or "").strip(),  # Col I
                puesto=puesto_val,
                codigo_cr=ceco_val,
                puesto_jefe_directo=str(vals[11] or "").strip(),  # Col L
                nombre_jefe_directo=str(vals[12] or "").strip(),  # Col M
                salario=_to_float(vals[13]),  # Col N
                categoria=str(vals[14] or "").strip(),  # Col O
                motivo_contrato=str(vals[15] or "").strip(),  # Col P
                tiempo_contrato=str(vals[16] or "").strip(),  # Col Q
                fecha_termino_contrato=_to_date(vals[17]),  # Col R
                modalidad_trabajo=str(vals[18] or "").strip(),  # Col S
                categoria_trabajador=str(vals[19] or "").strip(),  # Col T
                condicion_categoria=str(vals[20] or "").strip(),  # Col U
                tipo_ingreso=str(vals[21] or "").strip(),  # Col V
                codigo_reemplazo=str(vals[22] or "").strip(),  # Col W
                nombre_reemplazo=str(vals[23] or "").strip(),  # Col X
                cargo_reemplazo=str(vals[24] or "").strip(),  # Col Y
                fecha_cese_reemplazo=_to_date(vals[25]),  # Col Z
                estado=str(vals[26] or "IMPORTADO").strip(),  # Col AA
                unidad_funcional=str(vals[27] or "").strip() if len(vals) > 27 else "",
                area=str(vals[28] or "").strip() if len(vals) > 28 else "",
                plantilla_carta="CO BASE",
                incluye_eps=True,
            )
            db.add(s)
            count += 1

        # Insertar puestos únicos
        for puesto_nombre, ceco in puestos_set:
            if puesto_nombre and not db.query(Puesto).filter(Puesto.nombre == puesto_nombre).first():
                db.add(Puesto(nombre=puesto_nombre, codigo_ceco=ceco, activo=True))

        db.commit()
        print(f"  Solicitudes importadas: {count}")
        print(f"  Puestos únicos: {len(puestos_set)}")
    else:
        print(f"  No se encontró: {excel1}")

    # ── MOVIMIENTOS ORGANIZACIONALES ──
    excel2 = os.path.join(parent, "Estatus Cambios Organizacionales 2026-PROYECTO.xlsm")
    if os.path.exists(excel2):
        print(f"  Importando movimientos desde: {excel2}")
        wb2 = openpyxl.load_workbook(excel2, data_only=True, keep_vba=True)
        ws2 = wb2["CONSOLIDADO TOTAL"]
        count = 0

        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=29, values_only=True):
            nombre = row[5]  # Col F = NOMBRE
            if not nombre or str(nombre).strip() == "":
                continue

            m = Movimiento(
                periodo_pago=_to_date(row[0]),  # Col A
                tipo=str(row[1] or "").strip(),  # Col B
                fecha_cambio=_to_date(row[2]),  # Col C
                sucursal=str(row[3] or "").strip(),  # Col D
                codigo_colaborador=str(row[4] or "").strip(),  # Col E
                nombre=str(nombre).strip(),
                cargo_actual=str(row[6] or "").strip(),  # Col G
                fecha_ingreso=str(row[7] or "").strip(),  # Col H
                ceco_actual=str(row[8] or "").strip(),  # Col I
                jefe_directo_actual=str(row[9] or "").strip(),  # Col J
                categoria=str(row[10] or "").strip(),  # Col K
                sede_trabajo=str(row[11] or "").strip(),  # Col L
                departamento_actual=str(row[12] or "").strip(),  # Col M
                area_actual=str(row[13] or "").strip(),  # Col N
                seccion_actual=str(row[14] or "").strip(),  # Col O
                tipo_documento=str(row[15] or "").strip(),  # Col P
                cambio_cargo=str(row[16] or "").strip(),  # Col Q
                remuneracion=_to_float(row[17]),  # Col R
                bono=_to_float(row[18]),  # Col S
                movilidad=_to_float(row[19]),  # Col T
                fecha_inicio=_to_date(row[20]),  # Col U
                fecha_fin=_to_date(row[21]),  # Col V
                adryan=str(row[22] or "").strip(),  # Col W
                cargo_nuevo=str(row[23] or "").strip(),  # Col X
                ceco_nuevo=str(row[24] or "").strip(),  # Col Y
                sede_trabajo_nuevo=str(row[25] or "").strip() if len(row) > 25 else "",
                sucursal_nueva=str(row[26] or "").strip() if len(row) > 26 else "",
                categoria_nueva=str(row[27] or "").strip() if len(row) > 27 else "",
                jefe_directo_nuevo=str(row[28] or "").strip() if len(row) > 28 else "",
                estado="IMPORTADO",
            )
            db.add(m)
            count += 1

        db.commit()
        print(f"  Movimientos importados: {count}")
    else:
        print(f"  No se encontro: {excel2}")


def main():
    print("=" * 60)
    print("HR Operations — Seed Data")
    print("=" * 60)

    init_db()
    db = SessionLocal()

    try:
        # Check if already seeded
        existing = db.query(UnidadNegocio).count()
        if existing > 0:
            print("  Base de datos ya tiene datos. Saltando seed.")
            print(f"  Unidades: {existing}")
            print(f"  Solicitudes: {db.query(SolicitudIngreso).count()}")
            print(f"  Movimientos: {db.query(Movimiento).count()}")
            return

        seed_catalogos(db)
        seed_from_excel(db)
        print("\nSeed completado exitosamente!")
    finally:
        db.close()


if __name__ == "__main__":
    main()
